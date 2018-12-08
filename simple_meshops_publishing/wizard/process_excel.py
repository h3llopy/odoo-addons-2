# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
from openerp import api, models, fields
import base64
import tempfile
import os
import logging

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except (ImportError, IOError) as err:
    _logger.debug(err)


class SimpleMeShopsPublishing(models.TransientModel):
    _name = "simple_meshops_publishing.process_excel"
    data = fields.Binary(
        string="Worksheet to process"
    )
    name = fields.Char(
        'File Name',
    )
    state = fields.Selection(
        [('process', 'Process'),  # load spreadsheet
         ('download', 'Download')],  # download spreadsheet
        default="process"
    )
    date_from = fields.Date(
        default=fields.date.today()
    )

    @api.multi
    def process_data(self, fp_name):
        CODE_COL = 1
        PRICE_COL = 2
        FIRST_ROW = 2

        product_obj = self.env['product.product']
        for reg in self:
            products = product_obj.search([
                ('meshops_code', '=', True),
                ('write_date', '>=', reg.date_from)])

            # open worksheet
            wb = openpyxl.load_workbook(filename=fp_name,
                                        read_only=False,
                                        data_only=True)
            sheet = wb.active
            row = FIRST_ROW
            for product in products:
                sheet.cell(column=CODE_COL,
                           row=row).value = product.default_code
                sheet.cell(column=PRICE_COL,
                           row=row).value = product.final_price
                row += 1
            wb.save(fp_name)

    @api.multi
    def process_spreadsheet(self):
        # mover la planilla a un temporario
        (fileno, fp_name) = tempfile.mkstemp('.xlsx', 'openerp_')
        filename = os.path.dirname(os.path.realpath(__file__))
        filename = filename.replace('wizard', 'data/meshop_prices.xlsx')

        with open(filename, "r") as worksheet:
            data = worksheet.read()
        with open(fp_name, "w") as worksheet:
            worksheet.write(data)

        # procesar la planilla
        self.process_data(fp_name)

        # preparar para download
        with open(fp_name, "r") as worksheet:
            data = worksheet.read()

        for rec in self:
            rec.data = base64.encodestring(data)
            rec.state = 'download'
            rec.name = 'planilla.xlsx'
